"""
Converts output/jobs.json and output/skipped_companies.json into a single,
formatted Excel workbook: output/job_postings.xlsx

Two tabs:
  - "Job Postings": every job found, one row per posting
  - "Needs Manual Check": companies the scraper couldn't cover automatically
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
BODY_FONT = Font(name="Arial")
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_sheet(ws, headers, rows, col_widths):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(row)

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook(jobs_path="output/jobs.json",
                    skipped_path="output/skipped_companies.json",
                    output_path="output/job_postings.xlsx"):
    with open(jobs_path) as f:
        jobs = json.load(f)
    with open(skipped_path) as f:
        skipped = json.load(f)

    wb = Workbook()

    score_headers = ["New?", "Score", "Company", "Description", "Title", "Location",
                      "Reasoning", "Ambiguity Note", "IC Role?", "ATS", "URL"]
    col_widths = [6, 7, 22, 30, 36, 18, 42, 40, 9, 12, 42]

    def job_row(j):
        return ["Yes" if j.get("is_new") else "", j.get("score"), j.get("company"),
                 j.get("company_description"), j.get("title"), j.get("location"),
                 j.get("reasoning"), j.get("ambiguity_note"),
                 "Yes" if j.get("ic_role_flag") else "", j.get("source_ats"), j.get("url")]

    scored_jobs = sorted(
        [j for j in jobs if j.get("routing") == "Scored"],
        key=lambda j: (j.get("score") or 0), reverse=True
    )
    needs_review_jobs = [j for j in jobs if j.get("routing") == "Needs review"]
    non_fit_jobs = [j for j in jobs if j.get("routing") == "Non-fit"]
    new_jobs = sorted(
        [j for j in jobs if j.get("is_new")],
        key=lambda j: (j.get("score") or 0), reverse=True
    )

    ws_new = wb.active
    ws_new.title = "New This Run"
    _style_sheet(ws_new, score_headers, [job_row(j) for j in new_jobs], col_widths)

    ws_scored = wb.create_sheet("Scored List")
    _style_sheet(ws_scored, score_headers, [job_row(j) for j in scored_jobs], col_widths)

    ws_review = wb.create_sheet("Needs Review")
    _style_sheet(ws_review, score_headers, [job_row(j) for j in needs_review_jobs], col_widths)

    ws_nonfit = wb.create_sheet("Non-fits")
    _style_sheet(ws_nonfit, score_headers, [job_row(j) for j in non_fit_jobs], col_widths)

    ws_skipped = wb.create_sheet("Needs Manual Check")
    skip_headers = ["Company", "Description", "ATS Platform (as recorded)"]
    skip_rows = [[s.get("company"), s.get("sector"), s.get("ats_platform")] for s in skipped]
    _style_sheet(ws_skipped, skip_headers, skip_rows, col_widths=[30, 40, 32])

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = build_workbook()
    print(f"Wrote {path}")
